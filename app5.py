import openpyxl

wb = openpyxl.load_workbook("files/new_wb3.xlsx")

ws = wb.active

ws["B1"].value = 5
ws["C1"].value = 10
ws["D1"].value = "=B1*C1"

print(ws["D1"].value)
wb.save("files/new_wb3.xlsx")
